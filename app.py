"""
Application Streamlit — Système Intelligent de Traitement des Factures (Batch)
Traitement parallèle de plusieurs factures avec validation humaine groupée à chaque étape.
"""
import csv
import io
import tempfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

import streamlit as st

st.set_page_config(
    page_title="Système de Traitement des Factures",
    page_icon="🧾",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
    .step-header {
        font-size: 1.3rem; font-weight: bold; color: #1a5276;
        border-left: 5px solid #2e86c1; padding-left: 0.8rem;
        margin: 1.2rem 0 0.8rem 0;
    }
    .warn-box  { background:#fef3cd; border-left:4px solid #e6a817; padding:0.7rem 1rem; margin:0.4rem 0; border-radius:4px; color:#5a3e00; font-size:0.95rem; }
    .error-box { background:#fde8e8; border-left:4px solid #c0392b; padding:0.7rem 1rem; margin:0.4rem 0; border-radius:4px; color:#6b0f0f; font-size:0.95rem; }
    .info-box  { background:#ddeeff; border-left:4px solid #2980b9; padding:0.7rem 1rem; margin:0.4rem 0; border-radius:4px; color:#0a2a4a; font-size:0.95rem; }
    .warn-box b, .error-box b, .info-box b { font-weight: 700; }
</style>
""", unsafe_allow_html=True)

# ─── Session state ─────────────────────────────────────────────────────────────

def init_session():
    defaults = {"etape": 1, "batch": []}
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


def avancer(n: int):
    st.session_state.etape = n
    st.rerun()


def score_badge(score: float) -> str:
    if score >= 0.8:
        return f"🟢 {score:.0%}"
    if score >= 0.6:
        return f"🟡 {score:.0%}"
    return f"🔴 {score:.0%}"


# ─── Sidebar ──────────────────────────────────────────────────────────────────

def sidebar():
    etape = st.session_state.etape
    batch = st.session_state.batch

    st.sidebar.markdown("## 📊 Progression")
    etapes = [
        (1, "📁 Téléversement & Extraction"),
        (2, "📋 Validation des données"),
        (3, "🔍 Conformité fiscale"),
        (4, "📂 Classification PCM"),
        (5, "📊 Journal & Export"),
    ]
    for num, nom in etapes:
        if num < etape:
            st.sidebar.markdown(f"✅ ~~{nom}~~")
        elif num == etape:
            st.sidebar.markdown(f"**▶️ {nom}**")
        else:
            st.sidebar.markdown(f"⏳ {nom}")

    if batch:
        st.sidebar.markdown("---")
        st.sidebar.markdown(f"**{len(batch)} facture(s)** en cours")

    st.sidebar.markdown("---")
    if st.sidebar.button("🔄 Nouvelle analyse", use_container_width=True):
        for k in list(st.session_state.keys()):
            del st.session_state[k]
        st.rerun()


# ─── Workers parallèles ────────────────────────────────────────────────────────

def _new_state(nom: str, path: str, ext: str) -> dict:
    """Initialise un InvoiceState compatible avec le graphe LangGraph."""
    return {
        # Champs InvoiceState (graph/invoice_state.py)
        "fichier_nom": nom,
        "fichier_path": path,
        "fichier_type": ext,
        "statut": "fichier_reçu",
        "etape_courante": "Initialisation",
        "erreur": None,
        "donnees_extraites": None,
        "validation_extraction": None,
        "resultat_conformite": None,
        "validation_conformite": None,
        "code_comptable": None,
        "validation_classification": None,
        "ecriture_comptable": None,
        "rapport_final": None,
        "rapport_genere": False,
        "contexte_rag": [],
        "historique_validations": [],
        # Suivi d'erreurs par étape (UI batch)
        "erreur_conformite": None,
        "erreur_classification": None,
        "erreur_journal": None,
    }


def _run_parallel(fn, indices: list, label: str):
    """Exécute fn(batch[i]) en parallèle via ThreadPoolExecutor."""
    batch = st.session_state.batch
    if not indices:
        return
    bar = st.progress(0.0, text=label)
    results: dict = {}
    with ThreadPoolExecutor(max_workers=min(5, len(indices))) as executor:
        futures = {executor.submit(fn, batch[i]): i for i in indices}
        done = 0
        for future in as_completed(futures):
            idx = futures[future]
            results[idx] = future.result()
            done += 1
            bar.progress(done / len(indices), text=f"{label} ({done}/{len(indices)})")
    for idx, result in results.items():
        batch[idx] = result
    bar.empty()


# ── Workers : appellent les nœuds LangGraph (graph/nodes.py) ──────────────────
# Le mode batch parallèle exécute chaque nœud indépendamment par facture.
# Le graphe compilé (invoice_graph) gère le flux complet (voir demo_graph.py).

def _worker_extract(state: dict) -> dict:
    from graph.nodes import node_extract_invoice
    result = node_extract_invoice(state)
    if result.get("erreur"):
        return {**result, "donnees_extraites": None}
    return result


def _worker_conformite(state: dict) -> dict:
    from graph.nodes import node_check_compliance
    result = node_check_compliance(state)
    if result.get("erreur"):
        return {**result, "resultat_conformite": None,
                "erreur_conformite": result["erreur"], "erreur": None}
    return result


def _worker_classifier(state: dict) -> dict:
    from graph.nodes import node_classify_accounting
    result = node_classify_accounting(state)
    if result.get("erreur"):
        return {**result, "code_comptable": None,
                "erreur_classification": result["erreur"], "erreur": None}
    return result


def _worker_journal(state: dict) -> dict:
    from graph.nodes import node_generate_journal_entry
    result = node_generate_journal_entry(state)
    if result.get("erreur"):
        return {**result, "ecriture_comptable": None,
                "erreur_journal": result["erreur"], "erreur": None}
    return result


# ─── Étape 1 : Téléversement & extraction parallèle ──────────────────────────

def etape_1():
    st.markdown('<div class="step-header">📁 Étape 1 — Téléversement & Extraction</div>', unsafe_allow_html=True)
    st.info("ℹ️ Téléversez une ou plusieurs factures. L'extraction se fait en parallèle.")

    fichiers = st.file_uploader(
        "Téléverser des factures",
        type=["pdf", "xml", "txt", "png", "jpg", "jpeg", "webp"],
        accept_multiple_files=True,
        help="PDF natif, PDF scanné (OCR Vision), image PNG/JPG, XML ou texte",
    )
    if not fichiers:
        return

    st.success(f"✅ {len(fichiers)} fichier(s) sélectionné(s)")
    for f in fichiers:
        ext = Path(f.name).suffix.upper().lstrip(".")
        st.markdown(f"- **{f.name}** `{ext}`")

    if st.button("🚀 Lancer l'extraction de toutes les factures", type="primary"):
        states = []
        for f in fichiers:
            ext = Path(f.name).suffix.lower().lstrip(".")
            with tempfile.NamedTemporaryFile(delete=False, suffix=f".{ext}") as tmp:
                tmp.write(f.read())
            states.append(_new_state(f.name, tmp.name, ext))

        st.session_state.batch = states
        _run_parallel(_worker_extract, list(range(len(states))), "⚙️ Extraction en cours")
        avancer(2)


# ─── Étape 2 : Validation groupée des données extraites ───────────────────────

def etape_2():
    st.markdown('<div class="step-header">📋 Étape 2 — Validation des Données Extraites</div>', unsafe_allow_html=True)
    batch = st.session_state.batch

    ok = sum(1 for s in batch if s["donnees_extraites"])
    ko = len(batch) - ok
    c1, c2, c3 = st.columns(3)
    c1.metric("Factures", len(batch))
    c2.metric("Extraites avec succès", ok)
    c3.metric("Erreurs", ko)

    for i, state in enumerate(batch):
        inv = state["donnees_extraites"]
        icon = "✅" if inv else "❌"
        with st.expander(f"{icon} {i + 1}. {state['fichier_nom']}", expanded=not inv):
            if state["erreur"]:
                st.markdown(
                    f'<div class="error-box">❌ <b>Erreur d\'extraction :</b> {state["erreur"]}</div>',
                    unsafe_allow_html=True,
                )
                continue

            c1, c2 = st.columns([1, 3])
            c1.metric("Confiance", score_badge(inv.score_confiance))
            if inv.champs_manquants:
                c2.warning(f"Champs absents : {', '.join(inv.champs_manquants)}")

            col1, col2 = st.columns(2)
            with col1:
                st.markdown("**Fournisseur**")
                st.table({
                    "Champ": ["Nom", "N° Facture", "Date", "ICE", "IF", "Client"],
                    "Valeur": [
                        inv.fournisseur, inv.numero_facture, inv.date_facture,
                        inv.ice or "—", inv.if_fournisseur or "—",
                        inv.nom_client or "—",
                    ],
                })
            with col2:
                st.markdown("**Montants**")
                st.table({
                    "Champ": ["HT", "Taux TVA", "Montant TVA", "TTC"],
                    "Valeur": [
                        f"{inv.montant_ht:,.2f} MAD", f"{inv.taux_tva}%",
                        f"{inv.montant_tva:,.2f} MAD", f"{inv.montant_ttc:,.2f} MAD",
                    ],
                })

            if inv.lignes_facture:
                st.dataframe(
                    [{
                        "Description": l.description,
                        "Qté": l.quantite,
                        "PU HT": f"{l.prix_unitaire:,.2f}",
                        "Total HT": f"{l.montant_ht:,.2f}",
                        "TVA%": f"{l.taux_tva}%",
                    } for l in inv.lignes_facture],
                    use_container_width=True,
                )

            with st.expander("✏️ Corriger les montants"):
                ht = st.number_input("Montant HT",  value=inv.montant_ht,  format="%.2f", key=f"ht_{i}")
                tv = st.number_input("Montant TVA", value=inv.montant_tva, format="%.2f", key=f"tv_{i}")
                tc = st.number_input("Montant TTC", value=inv.montant_ttc, format="%.2f", key=f"tc_{i}")
                if st.button("💾 Enregistrer la correction", key=f"save_{i}"):
                    batch[i]["donnees_extraites"] = inv.model_copy(
                        update={"montant_ht": ht, "montant_tva": tv, "montant_ttc": tc}
                    )
                    st.success("✅ Correction enregistrée.")

    st.markdown("---")
    st.markdown("### ✅ Validation Humaine #1 — Confirmer les extractions")
    col1, col2 = st.columns(2)

    with col1:
        if st.button("✅ Accepter toutes et vérifier la conformité", type="primary", use_container_width=True):
            for s in batch:
                if s["donnees_extraites"] and not s["validation_extraction"]:
                    s["validation_extraction"] = "accepté"
            indices = [i for i, s in enumerate(batch) if s["validation_extraction"] == "accepté"]
            _run_parallel(_worker_conformite, indices, "🔍 Vérification de conformité")
            avancer(3)

    with col2:
        if st.button("❌ Tout annuler", use_container_width=True):
            for k in list(st.session_state.keys()):
                del st.session_state[k]
            st.rerun()


# ─── Étape 3 : Conformité groupée ─────────────────────────────────────────────

def etape_3():
    st.markdown('<div class="step-header">🔍 Étape 3 — Vérification de Conformité</div>', unsafe_allow_html=True)
    batch = st.session_state.batch
    actifs = [(i, s) for i, s in enumerate(batch) if s["validation_extraction"] == "accepté"]

    conformes = sum(1 for _, s in actifs if s.get("resultat_conformite") and s["resultat_conformite"].conforme)
    c1, c2, c3 = st.columns(3)
    c1.metric("Factures vérifiées", len(actifs))
    c2.metric("Conformes ✅", conformes)
    c3.metric("Non conformes ⚠️", len(actifs) - conformes)

    for i, state in actifs:
        compliance = state.get("resultat_conformite")
        if state.get("erreur_conformite"):
            with st.expander(f"❌ {state['fichier_nom']}"):
                st.error(state["erreur_conformite"])
            continue

        icon = "✅" if compliance.conforme else "⚠️"
        label = "CONFORME" if compliance.conforme else "NON CONFORME"
        with st.expander(
            f"{icon} {state['fichier_nom']} — {label} ({compliance.score_conformite:.0%})",
            expanded=not compliance.conforme,
        ):
            col1, col2 = st.columns(2)
            col1.metric("Statut", f"{icon} {label}")
            col2.metric("Score", score_badge(compliance.score_conformite))

            if compliance.avertissements:
                for a in compliance.avertissements:
                    niveau = str(getattr(a, "niveau", "mineur"))
                    css = {"critique": "error-box", "majeur": "warn-box", "mineur": "info-box"}.get(niveau, "info-box")
                    ref = f" — <i>{a.reference_legale}</i>" if a.reference_legale else ""
                    st.markdown(
                        f'<div class="{css}">⚠️ <b>[{niveau.upper()}]</b> {a.message}{ref}</div>',
                        unsafe_allow_html=True,
                    )
            else:
                st.success("✅ Aucun problème détecté — facture conforme.")

            if compliance.references_rag:
                with st.expander("📚 Références RAG consultées"):
                    for r in compliance.references_rag:
                        st.markdown(f"- {r}")

    st.markdown("---")
    st.markdown("### ✅ Validation Humaine #2 — Conformité fiscale")
    col1, col2 = st.columns(2)

    with col1:
        if st.button("✅ Confirmer et lancer la classification", type="primary", use_container_width=True):
            for _, s in actifs:
                if s.get("resultat_conformite"):
                    s["validation_conformite"] = "confirmé"
            indices = [i for i, s in actifs if s.get("validation_conformite") == "confirmé"]
            _run_parallel(_worker_classifier, indices, "📂 Classification comptable (PCM)")
            avancer(4)

    with col2:
        if st.button("⚠️ Ignorer les avertissements et continuer", use_container_width=True):
            for _, s in actifs:
                if s.get("resultat_conformite"):
                    s["validation_conformite"] = "ignoré"
            indices = [i for i, s in actifs if s.get("validation_conformite") == "ignoré"]
            _run_parallel(_worker_classifier, indices, "📂 Classification comptable (PCM)")
            avancer(4)


# ─── Étape 4 : Classification groupée ────────────────────────────────────────

def etape_4():
    st.markdown('<div class="step-header">📂 Étape 4 — Classification Comptable (PCM)</div>', unsafe_allow_html=True)
    batch = st.session_state.batch
    actifs = [(i, s) for i, s in enumerate(batch) if s.get("validation_conformite") in ("confirmé", "ignoré")]

    for i, state in actifs:
        code = state.get("code_comptable")
        if state.get("erreur_classification"):
            with st.expander(f"❌ {state['fichier_nom']}"):
                st.error(state["erreur_classification"])
            continue

        icon = "⚠️" if code.validation_humaine_requise else "📂"
        with st.expander(
            f"{icon} {state['fichier_nom']} — {code.code_comptable} | {code.libelle_compte}",
            expanded=code.validation_humaine_requise,
        ):
            col1, col2, col3 = st.columns(3)
            col1.metric("Code PCM", code.code_comptable)
            col2.metric("Libellé", code.libelle_compte)
            col3.metric("Confiance", score_badge(code.score_confiance))
            st.markdown(f"**Justification :** {code.justification}")

            if code.validation_humaine_requise:
                st.markdown(
                    '<div class="warn-box">⚠️ <b>Confiance insuffisante</b> — vérification recommandée</div>',
                    unsafe_allow_html=True,
                )

            if code.alternatives:
                with st.expander("📋 Codes alternatifs"):
                    for alt in code.alternatives:
                        st.markdown(f"- `{alt.code}` — {alt.libelle} ({alt.score:.0%})")

            with st.expander("✏️ Corriger ce code"):
                code_c = st.text_input("Code PCM", value=code.code_comptable, key=f"cc_{i}")
                lib_c  = st.text_input("Libellé",  value=code.libelle_compte,  key=f"cl_{i}")
                if st.button("💾 Appliquer la correction", key=f"apply_{i}"):
                    batch[i]["code_comptable"] = code.model_copy(
                        update={"code_comptable": code_c, "libelle_compte": lib_c}
                    )
                    batch[i]["validation_classification"] = "corrigé"
                    st.success("✅ Correction enregistrée.")

    st.markdown("---")
    st.markdown("### ✅ Validation Humaine #3 — Codes comptables")

    if st.button("✅ Accepter et générer les écritures comptables", type="primary", use_container_width=True):
        for i, s in actifs:
            if not s.get("validation_classification"):
                s["validation_classification"] = "accepté"
        indices = [
            i for i, s in actifs
            if s.get("validation_classification") in ("accepté", "corrigé")
        ]
        _run_parallel(_worker_journal, indices, "📝 Génération des écritures comptables")
        avancer(5)


# ─── Étape 5 : Journal comptable & Export XLSX/CSV ───────────────────────────

def _parse_date(date_str: str):
    """Parse DD/MM/YYYY for sorting; unknown formats sort last."""
    try:
        return datetime.strptime(date_str, "%d/%m/%Y")
    except Exception:
        return datetime.max


def _build_journal_rows(batch: list) -> list:
    rows = []
    for state in batch:
        entry = state.get("ecriture_comptable")
        if not entry:
            continue
        inv = state["donnees_extraites"]
        conforme = (
            state.get("resultat_conformite") and state["resultat_conformite"].conforme
        )
        for ligne in entry.ecritures:
            rows.append({
                "Date": entry.date_ecriture,
                "Référence": entry.reference,
                "N° Facture": inv.numero_facture,
                "Fournisseur": inv.fournisseur,
                "Sens": str(ligne.sens),
                "Compte": ligne.compte,
                "Libellé": ligne.libelle,
                "Montant (MAD)": ligne.montant,
                "Équilibré": "Oui" if entry.equilibre else "Non",
                "Conforme": "Oui" if conforme else "Non",
            })
    rows.sort(key=lambda r: _parse_date(r["Date"]))
    return rows


def _export_csv(rows: list) -> bytes:
    buf = io.StringIO()
    if rows:
        writer = csv.DictWriter(buf, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)
    return buf.getvalue().encode("utf-8-sig")


def _export_xlsx(rows: list) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Journal Comptable"

    if not rows:
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    headers = list(rows[0].keys())

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor="1A5276")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[1].height = 22

    debit_fill  = PatternFill("solid", fgColor="FDECEA")
    credit_fill = PatternFill("solid", fgColor="EAFAF1")

    for row_idx, row in enumerate(rows, 2):
        fill = debit_fill if "bit" in str(row.get("Sens", "")).lower() else credit_fill
        for col, key in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col, value=row[key])
            cell.border = border
            cell.fill = fill
            if key == "Montant (MAD)":
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")

    for col in ws.columns:
        width = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(width + 4, 45)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def etape_5():
    st.markdown('<div class="step-header">📊 Étape 5 — Journal Comptable & Export</div>', unsafe_allow_html=True)
    batch = st.session_state.batch
    avec_ecriture = [s for s in batch if s.get("ecriture_comptable")]

    if not avec_ecriture:
        st.error("Aucune écriture comptable disponible.")
        return

    st.markdown(
        '<div class="warn-box">⚠️ <b>Important :</b> Ces écritures sont des <b>suggestions IA</b>. '
        "Elles doivent être validées par un comptable qualifié avant tout enregistrement.</div>",
        unsafe_allow_html=True,
    )

    rows = _build_journal_rows(avec_ecriture)

    st.markdown("### 📒 Journal des écritures — toutes factures")
    st.dataframe(rows, use_container_width=True)

    st.markdown("### 📋 Détail par facture")
    for state in avec_ecriture:
        entry = state["ecriture_comptable"]
        inv   = state["donnees_extraites"]
        eq    = "✅ Équilibrée" if entry.equilibre else "❌ Déséquilibrée"
        total_d = sum(l.montant for l in entry.ecritures if "bit" in str(l.sens).lower())
        total_c = sum(l.montant for l in entry.ecritures if "dit" in str(l.sens).lower())

        with st.expander(
            f"🧾 {state['fichier_nom']} | {inv.fournisseur} | TTC : {inv.montant_ttc:,.2f} MAD | {eq}"
        ):
            st.dataframe(
                [{
                    "Sens": str(l.sens),
                    "Compte": l.compte,
                    "Libellé": l.libelle,
                    "Montant (MAD)": f"{l.montant:,.2f}",
                } for l in entry.ecritures],
                use_container_width=True,
            )
            col1, col2 = st.columns(2)
            col1.metric("Total Débit",  f"{total_d:,.2f} MAD")
            col2.metric("Total Crédit", f"{total_c:,.2f} MAD")
            if entry.avertissement:
                st.markdown(
                    f'<div class="warn-box">⚠️ {entry.avertissement}</div>',
                    unsafe_allow_html=True,
                )

    st.markdown("---")
    st.markdown("### ⬇️ Exporter le journal")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    col1, col2 = st.columns(2)

    with col1:
        st.download_button(
            label="⬇️ Télécharger CSV",
            data=_export_csv(rows),
            file_name=f"journal_comptable_{ts}.csv",
            mime="text/csv",
            use_container_width=True,
            type="primary",
        )
    with col2:
        st.download_button(
            label="⬇️ Télécharger Excel (XLSX)",
            data=_export_xlsx(rows),
            file_name=f"journal_comptable_{ts}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    st.markdown("---")
    n_ok  = sum(1 for s in avec_ecriture if s["ecriture_comptable"].equilibre)
    n_ttc = sum(s["donnees_extraites"].montant_ttc for s in avec_ecriture)
    col1, col2, col3 = st.columns(3)
    col1.metric("Factures traitées", len(avec_ecriture))
    col2.metric("Écritures équilibrées", n_ok)
    col3.metric("Total TTC traité", f"{n_ttc:,.2f} MAD")


# ─── Main ──────────────────────────────────────────────────────────────────────

def main():
    init_session()
    sidebar()
    st.markdown(
        '<h1 style="color:#1a5276;">🧾 Système Intelligent de Traitement des Factures</h1>',
        unsafe_allow_html=True,
    )
    st.caption("Traitement batch parallèle · Conformité fiscale marocaine · Classification PCM · Export XLSX/CSV")
    st.divider()

    dispatch = {1: etape_1, 2: etape_2, 3: etape_3, 4: etape_4, 5: etape_5}
    dispatch.get(st.session_state.etape, etape_1)()


if __name__ == "__main__":
    main()
